import os
import email
import tempfile
import pytesseract
from bs4 import BeautifulSoup
from email import policy
from email.parser import BytesParser
from email.utils import getaddresses
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

def parse_eml_file(file):
    msg = BytesParser(policy=policy.default).parse(file)

    parsed = {
        "from_address": "",
        "to_address": "",
        "subject": msg.get("Subject", ""),
        "date": msg.get("Date", ""),
        "text": "",
        "html": "",
        "attachments": [],
        "embedded_images": [],
        "attachment_text_summary": "",
        "embedded_image_text": ""
    }

    from_emails = getaddresses([msg.get("From", "")])
    to_emails = getaddresses([msg.get("To", "")])
    parsed["from_address"] = from_emails[0][1] if from_emails else ""
    parsed["to_address"] = to_emails[0][1] if to_emails else ""

    attachment_texts = []
    image_ocr_texts = []

    for part in msg.walk():
        content_type = part.get_content_type()
        content_disposition = part.get_content_disposition()
        filename = part.get_filename()

        if content_type == "text/plain" and not parsed["text"]:
            parsed["text"] = part.get_content().strip()
        elif content_type == "text/html" and not parsed["html"]:
            html = part.get_content()
            parsed["html"] = html
            if not parsed["text"]:
                soup = BeautifulSoup(html, "html5lib")
                parsed["text"] = soup.get_text().strip()
        elif content_disposition == "attachment" and filename:
            print(f"📎 Found attachment: {filename} ({content_type})")
            payload = part.get_payload(decode=True)
            parsed["attachments"].append({"filename": filename, "content_type": content_type})
            text = extract_text_from_known_types(filename, payload)
            if text:
                attachment_texts.append(f"[{filename}]\n{text}")
        elif content_type.startswith("image/") and content_disposition != "attachment":
            payload = part.get_payload(decode=True)
            parsed["embedded_images"].append(content_type)
            ocr_text = extract_text_from_image(payload)
            if ocr_text:
                image_ocr_texts.append(ocr_text)

    parsed["attachment_text_summary"] = "\n\n".join(attachment_texts)
    parsed["embedded_image_text"] = "\n\n".join(image_ocr_texts)

    print(f"✅ From: {parsed['from_address']}")
    print(f"✅ To: {parsed['to_address']}")

    return parsed

def extract_text_from_known_types(filename, payload):
    import subprocess
    ext = os.path.splitext(filename)[1].lower()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(payload)
            tmp.flush()
            print(f"🔍 Attempting to extract from: {filename}")
            
            if ext == ".pdf":
                reader = PdfReader(tmp.name)
                extracted = "\n".join(page.extract_text() or "" for page in reader.pages)

                if extracted.strip():
                    print(f"✅ Extracted text from PDF: {filename}")
                    return extracted
                else:
                    print(f"⚠️ No text in PDF. Trying OCR fallback: {filename}")
                    from pdf2image import convert_from_path
                    images = convert_from_path(tmp.name)
                    return "\n".join(pytesseract.image_to_string(img) for img in images)

            elif ext == ".docx":
                return "\n".join(p.text for p in Document(tmp.name).paragraphs)

            elif ext == ".xlsx":
                wb = load_workbook(tmp.name, read_only=True)
                return "\n".join(
                    str(cell.value)
                    for sheet in wb
                    for row in sheet.iter_rows()
                    for cell in row if cell.value
                )

            elif ext in [".txt", ".csv"]:
                return open(tmp.name, "r", encoding="utf-8", errors="ignore").read()
    except Exception as e:
        print(f"⚠️ Could not extract text from {filename}: {e}")
    return ""

def extract_text_from_image(image_bytes):
    tess_cmd = os.getenv("TESSERACT_CMD")
    if not tess_cmd:
        print("⚠️ Warning: TESSERACT_CMD not set — OCR will not work")
        return ""

    try:
        pytesseract.pytesseract.tesseract_cmd = tess_cmd
        return pytesseract.image_to_string(Image.open(BytesIO(image_bytes)))
    except Exception as e:
        print(f"⚠️ OCR failed: {e}")
        return ""
